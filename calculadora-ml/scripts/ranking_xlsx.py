# -*- coding: utf-8 -*-
"""
Gera a planilha RANKING_ANUNCIAR_ML.xlsx a partir do ranking-anunciar-ml.json
(saida do scripts/mapa-calor.mjs). Duas abas: a lista curta pra comecar e o
ranking completo, ambas com filtro e cores por acao recomendada.

Uso:  python scripts/ranking_xlsx.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(os.path.dirname(AQUI))  # pasta mercado-livre-ok
ENTRADA = os.path.join(RAIZ, "ranking-anunciar-ml.json")
SAIDA = os.path.join(RAIZ, "RANKING_ANUNCIAR_ML.xlsx")

CABECALHO = [
    ("#", 5), ("Score", 7), ("Acao", 16), ("Cod", 8), ("Produto", 52),
    ("Marca", 16), ("Grupo", 16), ("Custo R$", 11), ("Preco p/ competir R$", 13),
    ("Sobra R$", 11), ("Sobra %", 9), ("Visitas/dia", 11), ("Tendencia", 13),
    ("Concorrentes", 12), ("Oficiais", 9), ("Por que", 40), ("Link do anuncio", 22),
]

COR = {
    "ANUNCIAR JA": "C6EFCE",     # verde
    "SEGUNDA RODADA": "FFF2CC",  # amarelo
    "SO COM CUIDADO": "FCE4D6",  # laranja
    "NAO AGORA": "F2F2F2",       # cinza
}
CABECALHO_FILL = PatternFill("solid", fgColor="1B5E20")
CABECALHO_FONTE = Font(color="FFFFFF", bold=True, size=10)


def escrever_aba(wb, titulo, itens):
    ws = wb.create_sheet(titulo)
    ws.append([c for c, _ in CABECALHO])
    for i, (_, larg) in enumerate(CABECALHO, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg
        cel = ws.cell(row=1, column=i)
        cel.fill = CABECALHO_FILL
        cel.font = CABECALHO_FONTE
        cel.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for it in itens:
        ws.append([
            it["pos"], it["score"], it["acao"], it["cod"], it["produto"],
            it.get("marca") or "", it.get("grupo") or "",
            it.get("custo"), it.get("preco"), it.get("sobra_rs"),
            (it.get("sobra_pct") or 0) * 100,
            round(it.get("visitas_dia") or 0, 2), it.get("tendencia") or "",
            it.get("n_vend"), it.get("n_oficiais"),
            " · ".join(it.get("motivos") or []),
            it.get("url_conc") or it.get("url_cat") or "",
        ])
        linha = ws.max_row
        fill = COR.get(it["acao"])
        if fill:
            pf = PatternFill("solid", fgColor=fill)
            for col in range(1, 4):
                ws.cell(row=linha, column=col).fill = pf
        for col in (8, 9, 10):
            ws.cell(row=linha, column=col).number_format = 'R$ #,##0.00'
        ws.cell(row=linha, column=11).number_format = '0.0"%"'
        ws.cell(row=linha, column=12).number_format = '0.00'
        ws.cell(row=linha, column=5).alignment = Alignment(wrap_text=False)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CABECALHO))}{ws.max_row}"
    return ws


def main():
    d = json.load(open(ENTRADA, encoding="utf-8"))
    itens = d["itens"]
    comecar = [x for x in itens if x["acao"] in ("ANUNCIAR JA", "SEGUNDA RODADA")]

    wb = Workbook()
    wb.remove(wb.active)

    # capa curta explicando o criterio
    cp = wb.create_sheet("Como ler")
    cp.column_dimensions["A"].width = 110
    linhas = [
        ("COMO LER ESTE RANKING", True),
        ("", False),
        (f"Base: {d['total']} produtos da sua planilha que ja estavam casados com o catalogo do Mercado Livre.", False),
        (f"Precos e margem da analise de {d['data_analise_precos']}, revalidados com o preco de hoje no ML.", False),
        (f"Procura medida pelas visitas reais dos anuncios, janela de {d['janela_dias']} dias.", False),
        ("", False),
        ("ANUNCIAR JA     = sobra 15%+ igualando o mais barato, tem procura diaria, sem loja oficial, ate 15 concorrentes.", False),
        ("SEGUNDA RODADA  = sobra 12%+ e tem procura, mas esbarra em concorrencia ou loja oficial.", False),
        ("SO COM CUIDADO  = a conta fecha, mas a procura e fraca ou o cenario e dificil.", False),
        ("NAO AGORA       = margem abaixo de 12% ou quase ninguem procura.", False),
        ("", False),
        ("Score 0-100 = 40% procura (visitas/dia) + 35% margem + 25% pouca concorrencia.", False),
        ("", False),
        ("IMPORTANTE: visita nao e venda. A API do Mercado Livre nao abre quantidade vendida por anuncio.", True),
        ("O ranking diz onde vale gastar seu tempo anunciando, nao garante que vai vender.", False),
        ("Sua conta e nova: o mesmo anuncio vende menos que o de um MercadoLider. Comece pelos de menor risco.", False),
        ("Saldo de estoque NAO entra aqui (nao veio na planilha) - confira antes de anunciar.", False),
    ]
    for texto, negrito in linhas:
        cp.append([texto])
        if negrito:
            cp.cell(row=cp.max_row, column=1).font = Font(bold=True, size=12)

    escrever_aba(wb, "Comecar por estes", comecar)
    escrever_aba(wb, "Ranking completo", itens)
    wb.save(SAIDA)
    print(f"OK: {SAIDA}")
    print(f"  'Comecar por estes': {len(comecar)} produtos")
    print(f"  'Ranking completo':  {len(itens)} produtos")


if __name__ == "__main__":
    main()
